#!/usr/bin/env python3
"""
Financial Manager GUI - Bootstrap-style tkinter interface
Provides a modern GUI for managing financial data, importing CSVs, and updating balances.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import os
import csv
import threading
import subprocess
import time
from datetime import datetime
from pathlib import Path
import shutil
import sys

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

class FinancialManagerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Financial Manager - Kunz Family Finances")
        self.root.geometry("1200x800")

        # Base paths
        self.base_path = Path(__file__).parent
        self.archive_path = self.base_path / "Archive"
        self.processed_path = self.archive_path / "processed"
        self.raw_path = self.archive_path / "raw"
        self.scripts_path = self.base_path / "scripts"

        # Auto-rebuild watcher
        self.watcher_running = False
        self.watcher_thread = None

        # Setup styling
        self.setup_styles()

        # Create main layout
        self.create_widgets()

        # Load initial data
        self.load_financial_data()

    def setup_styles(self):
        """Setup bootstrap-style colors and themes"""
        style = ttk.Style()

        # Try to use 'clam' theme as base for better customization
        try:
            style.theme_use('clam')
        except:
            pass

        # Bootstrap-inspired colors
        self.colors = {
            'primary': '#007bff',
            'success': '#28a745',
            'danger': '#dc3545',
            'warning': '#ffc107',
            'info': '#17a2b8',
            'light': '#f8f9fa',
            'dark': '#343a40',
            'secondary': '#6c757d',
            'bg': '#ffffff',
            'text': '#212529'
        }

        # Configure styles
        style.configure('Title.TLabel', font=('Helvetica', 16, 'bold'), foreground=self.colors['dark'])
        style.configure('Heading.TLabel', font=('Helvetica', 12, 'bold'), foreground=self.colors['primary'])
        style.configure('Success.TLabel', foreground=self.colors['success'], font=('Helvetica', 10, 'bold'))
        style.configure('Danger.TLabel', foreground=self.colors['danger'], font=('Helvetica', 10, 'bold'))
        style.configure('Warning.TLabel', foreground=self.colors['warning'], font=('Helvetica', 10, 'bold'))

        style.configure('Primary.TButton', font=('Helvetica', 10, 'bold'))
        style.configure('Success.TButton', font=('Helvetica', 10, 'bold'))
        style.configure('Danger.TButton', font=('Helvetica', 10, 'bold'))

        # Card frame style
        style.configure('Card.TFrame', background='white', relief='raised')

    def create_widgets(self):
        """Create main GUI layout"""
        # Header
        header = ttk.Frame(self.root, style='Card.TFrame')
        header.pack(fill='x', padx=10, pady=10)

        title = ttk.Label(header, text="💰 Financial Manager", style='Title.TLabel')
        title.pack(side='left', padx=10, pady=10)

        # Status indicator
        self.status_label = ttk.Label(header, text="● Ready", foreground=self.colors['success'])
        self.status_label.pack(side='right', padx=10)

        # Auto-rebuild toggle
        self.auto_rebuild_var = tk.BooleanVar(value=False)
        auto_rebuild_check = ttk.Checkbutton(
            header,
            text="Auto-Rebuild",
            variable=self.auto_rebuild_var,
            command=self.toggle_auto_rebuild
        )
        auto_rebuild_check.pack(side='right', padx=10)

        # Main notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=5)

        # Create tabs
        self.create_dashboard_tab()
        self.create_import_tab()
        self.create_update_tab()
        self.create_comparison_tab()
        self.create_automation_tab()
        self.create_reports_tab()

        # Footer with action buttons
        footer = ttk.Frame(self.root)
        footer.pack(fill='x', padx=10, pady=10)

        ttk.Button(
            footer,
            text="🔄 Rebuild Dashboard",
            command=self.rebuild_dashboard,
            style='Primary.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            footer,
            text="📸 Save Snapshot",
            command=self.save_snapshot,
            style='Success.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            footer,
            text="🌐 Open Dashboard",
            command=self.open_dashboard,
            style='Primary.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            footer,
            text="❌ Exit",
            command=self.root.quit,
            style='Danger.TButton'
        ).pack(side='right', padx=5)

    def create_dashboard_tab(self):
        """Dashboard overview tab"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📊 Dashboard")

        # Create scrollable frame
        canvas = tk.Canvas(tab, bg='white')
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Title
        ttk.Label(scrollable_frame, text="Financial Overview", style='Title.TLabel').pack(pady=10)

        # Cards container
        cards_frame = ttk.Frame(scrollable_frame)
        cards_frame.pack(fill='both', expand=True, padx=20, pady=10)

        # Store card frames
        self.dashboard_cards = {}

        # Create metric cards
        metrics = [
            ('Liquid Cash', 'cash', 'success'),
            ('Total Debt', 'debt', 'danger'),
            ('Net Worth', 'networth', 'info'),
            ('Monthly Recurring', 'recurring', 'warning'),
            ('Monthly Income', 'income', 'success'),
            ('Emergency Fund Progress', 'emergency', 'primary')
        ]

        row = 0
        col = 0
        for title, key, color in metrics:
            card = self.create_metric_card(cards_frame, title, "$0", color)
            card.grid(row=row, column=col, padx=10, pady=10, sticky='ew')
            self.dashboard_cards[key] = card

            col += 1
            if col > 2:
                col = 0
                row += 1

        # Configure grid weights
        for i in range(3):
            cards_frame.columnconfigure(i, weight=1)

        # Alerts section
        alerts_frame = ttk.LabelFrame(scrollable_frame, text="⚠️ Alerts", padding=10)
        alerts_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.alerts_text = scrolledtext.ScrolledText(alerts_frame, height=6, wrap=tk.WORD)
        self.alerts_text.pack(fill='both', expand=True)

        # Recent activity
        activity_frame = ttk.LabelFrame(scrollable_frame, text="📝 Recent Activity", padding=10)
        activity_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.activity_text = scrolledtext.ScrolledText(activity_frame, height=8, wrap=tk.WORD)
        self.activity_text.pack(fill='both', expand=True)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def create_metric_card(self, parent, title, value, color):
        """Create a metric card widget"""
        card = ttk.Frame(parent, relief='raised', borderwidth=2)
        card.configure(padding=15)

        title_label = ttk.Label(card, text=title, font=('Helvetica', 10))
        title_label.pack()

        value_label = ttk.Label(
            card,
            text=value,
            font=('Helvetica', 20, 'bold'),
            foreground=self.colors.get(color, self.colors['dark'])
        )
        value_label.pack(pady=5)

        # Store value label for updates
        card.value_label = value_label

        return card

    def create_import_tab(self):
        """CSV Import tab"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📥 Import CSV")

        # Instructions
        instructions = ttk.LabelFrame(tab, text="Instructions", padding=10)
        instructions.pack(fill='x', padx=20, pady=10)

        ttk.Label(
            instructions,
            text="Import financial data from CSV or Excel files (supports multiple files).\n"
                 "• Supports: CSV (.csv), Excel (.xlsx, .xls)\n"
                 "• Excel files are automatically converted to CSV\n"
                 "• All files are validated before importing",
            justify='left'
        ).pack()

        # File selection
        file_frame = ttk.LabelFrame(tab, text="Select Files (CSV or Excel)", padding=10)
        file_frame.pack(fill='x', padx=20, pady=10)

        self.import_files = []  # Store multiple files
        self.import_file_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.import_file_var, width=60)
        file_entry.pack(side='left', padx=5, fill='x', expand=True)

        ttk.Button(
            file_frame,
            text="Browse...",
            command=self.browse_import_file
        ).pack(side='left', padx=5)

        ttk.Button(
            file_frame,
            text="Validate All",
            command=self.validate_files,
            style='Primary.TButton'
        ).pack(side='left', padx=5)

        # Preview area
        preview_frame = ttk.LabelFrame(tab, text="Preview & Validation", padding=10)
        preview_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.import_preview = scrolledtext.ScrolledText(preview_frame, height=15, wrap=tk.WORD)
        self.import_preview.pack(fill='both', expand=True)

        # Import button
        import_btn_frame = ttk.Frame(tab)
        import_btn_frame.pack(fill='x', padx=20, pady=10)

        self.import_btn = ttk.Button(
            import_btn_frame,
            text="✓ Import & Process All",
            command=self.import_csv,
            style='Success.TButton',
            state='disabled'
        )
        self.import_btn.pack(side='right')

        # Clear selection button
        ttk.Button(
            import_btn_frame,
            text="Clear Selection",
            command=self.clear_import_selection
        ).pack(side='right', padx=5)

    def create_update_tab(self):
        """Quick update forms tab"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="✏️ Update Balances")

        # Create scrollable frame
        canvas = tk.Canvas(tab, bg='white')
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Cash Accounts
        cash_frame = ttk.LabelFrame(scrollable_frame, text="💵 Cash Accounts", padding=10)
        cash_frame.pack(fill='x', padx=20, pady=10)

        self.cash_entries = {}

        # Debt Accounts
        debt_frame = ttk.LabelFrame(scrollable_frame, text="💳 Debt Balances", padding=10)
        debt_frame.pack(fill='x', padx=20, pady=10)

        self.debt_entries = {}

        # Recurring Expenses
        recurring_frame = ttk.LabelFrame(scrollable_frame, text="🔄 Recurring Expenses", padding=10)
        recurring_frame.pack(fill='x', padx=20, pady=10)

        self.recurring_entries = {}

        # Save button
        save_frame = ttk.Frame(scrollable_frame)
        save_frame.pack(fill='x', padx=20, pady=20)

        ttk.Button(
            save_frame,
            text="💾 Save All Changes",
            command=self.save_all_updates,
            style='Success.TButton'
        ).pack(side='right')

        ttk.Button(
            save_frame,
            text="🔄 Refresh from File",
            command=self.load_financial_data,
            style='Primary.TButton'
        ).pack(side='right', padx=5)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Store references
        self.cash_frame = cash_frame
        self.debt_frame = debt_frame
        self.recurring_frame = recurring_frame

    def create_comparison_tab(self):
        """Month-over-month comparison tab"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📈 Comparisons")

        ttk.Label(tab, text="Month-over-Month Analysis", style='Title.TLabel').pack(pady=10)

        # Controls
        control_frame = ttk.Frame(tab)
        control_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(control_frame, text="Compare:").pack(side='left', padx=5)

        self.compare_month1 = ttk.Combobox(control_frame, width=15)
        self.compare_month1.pack(side='left', padx=5)

        ttk.Label(control_frame, text="vs").pack(side='left', padx=5)

        self.compare_month2 = ttk.Combobox(control_frame, width=15)
        self.compare_month2.pack(side='left', padx=5)

        ttk.Button(
            control_frame,
            text="Compare",
            command=self.run_comparison,
            style='Primary.TButton'
        ).pack(side='left', padx=5)

        # Results area
        results_frame = ttk.LabelFrame(tab, text="Comparison Results", padding=10)
        results_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.comparison_text = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD)
        self.comparison_text.pack(fill='both', expand=True)

        # Load available snapshots
        self.load_snapshots()

    def create_automation_tab(self):
        """Automation settings tab"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="⚙️ Automation")

        ttk.Label(tab, text="Automation Settings", style='Title.TLabel').pack(pady=10)

        # Auto-rebuild section
        rebuild_frame = ttk.LabelFrame(tab, text="Auto-Rebuild Dashboard", padding=15)
        rebuild_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(
            rebuild_frame,
            text="Automatically rebuild the dashboard when financial data changes."
        ).pack(anchor='w')

        self.rebuild_status = ttk.Label(rebuild_frame, text="Status: Stopped", foreground='red')
        self.rebuild_status.pack(anchor='w', pady=5)

        btn_frame = ttk.Frame(rebuild_frame)
        btn_frame.pack(fill='x', pady=5)

        ttk.Button(
            btn_frame,
            text="▶️ Start Watcher",
            command=self.start_watcher,
            style='Success.TButton'
        ).pack(side='left', padx=5)

        ttk.Button(
            btn_frame,
            text="⏹️ Stop Watcher",
            command=self.stop_watcher,
            style='Danger.TButton'
        ).pack(side='left', padx=5)

        # Monthly snapshot section
        snapshot_frame = ttk.LabelFrame(tab, text="Monthly Snapshots", padding=15)
        snapshot_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(
            snapshot_frame,
            text="Automatically save monthly snapshots on the 1st of each month."
        ).pack(anchor='w')

        self.auto_snapshot_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            snapshot_frame,
            text="Enable automatic monthly snapshots",
            variable=self.auto_snapshot_var
        ).pack(anchor='w', pady=5)

        ttk.Button(
            snapshot_frame,
            text="💾 Save Snapshot Now",
            command=self.save_snapshot,
            style='Primary.TButton'
        ).pack(anchor='w', pady=5)

        # Data validation section
        validation_frame = ttk.LabelFrame(tab, text="Data Validation", padding=15)
        validation_frame.pack(fill='x', padx=20, pady=10)

        self.validate_on_save_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            validation_frame,
            text="Validate data before saving",
            variable=self.validate_on_save_var
        ).pack(anchor='w')

        self.alert_large_changes_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            validation_frame,
            text="Alert on large balance changes (>20%)",
            variable=self.alert_large_changes_var
        ).pack(anchor='w')

        ttk.Button(
            validation_frame,
            text="🔍 Run Validation Now",
            command=self.run_validation,
            style='Primary.TButton'
        ).pack(anchor='w', pady=5)

        # Watcher log
        log_frame = ttk.LabelFrame(tab, text="Activity Log", padding=10)
        log_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.automation_log = scrolledtext.ScrolledText(log_frame, height=10, wrap=tk.WORD)
        self.automation_log.pack(fill='both', expand=True)

    def create_reports_tab(self):
        """View reports tab"""
        tab = ttk.Frame(self.notebook)
        self.notebook.add(tab, text="📄 Reports")

        # Report selector
        selector_frame = ttk.Frame(tab)
        selector_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(selector_frame, text="Select Report:", style='Heading.TLabel').pack(side='left', padx=5)

        self.report_selector = ttk.Combobox(selector_frame, width=40)
        self.report_selector.pack(side='left', padx=5, fill='x', expand=True)
        self.report_selector.bind('<<ComboboxSelected>>', self.load_selected_report)

        ttk.Button(
            selector_frame,
            text="🔄 Refresh",
            command=self.refresh_reports_list
        ).pack(side='left', padx=5)

        # Report viewer
        viewer_frame = ttk.LabelFrame(tab, text="Report Content", padding=10)
        viewer_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.report_text = scrolledtext.ScrolledText(viewer_frame, wrap=tk.WORD)
        self.report_text.pack(fill='both', expand=True)

        # Load reports list
        self.refresh_reports_list()

    # ================== Data Loading ==================

    def flatten_nested_dict(self, nested_dict, value_key='balance'):
        """Flatten nested dictionary structure"""
        flat = {}
        for category, items in nested_dict.items():
            if isinstance(items, dict):
                # Check if this is a nested category structure
                has_nested_items = any(isinstance(v, dict) for v in items.values())

                if has_nested_items:
                    # This is a category with nested items
                    for item_key, item_val in items.items():
                        if isinstance(item_val, dict):
                            if value_key in item_val:
                                flat[item_key] = float(item_val[value_key])
                            elif 'amount' in item_val:
                                flat[item_key] = float(item_val['amount'])
                else:
                    # This item itself has the value
                    if value_key in items:
                        flat[category] = float(items[value_key])
                    elif 'amount' in items:
                        flat[category] = float(items['amount'])
            elif isinstance(items, (int, float)):
                flat[category] = float(items)

        return flat

    def load_financial_data(self):
        """Load all financial data from JSON files"""
        try:
            # Load financial config
            config_file = self.processed_path / "financial_config.json"
            if config_file.exists():
                with open(config_file, 'r') as f:
                    self.financial_config = json.load(f)
            else:
                self.financial_config = self.create_default_config()

            # Load dashboard data
            dashboard_file = self.processed_path / "dashboard_data.json"
            if dashboard_file.exists():
                with open(dashboard_file, 'r') as f:
                    self.dashboard_data = json.load(f)
            else:
                self.dashboard_data = {}

            # Update UI
            self.populate_update_forms()
            self.update_dashboard_metrics()
            self.log_activity("Financial data loaded successfully")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load financial data: {str(e)}")
            self.log_activity(f"ERROR: {str(e)}")

    def create_default_config(self):
        """Create default financial config structure"""
        return {
            "cash_accounts": {},
            "debt_balances": {},
            "recurring_expenses": {},
            "metadata": {
                "last_updated": datetime.now().isoformat(),
                "version": "1.0"
            }
        }

    def populate_update_forms(self):
        """Populate the update forms with current data"""
        # Clear existing entries
        for widget in self.cash_frame.winfo_children():
            if isinstance(widget, ttk.Frame):
                widget.destroy()

        for widget in self.debt_frame.winfo_children():
            if isinstance(widget, ttk.Frame):
                widget.destroy()

        for widget in self.recurring_frame.winfo_children():
            if isinstance(widget, ttk.Frame):
                widget.destroy()

        self.cash_entries.clear()
        self.debt_entries.clear()
        self.recurring_entries.clear()

        # Populate cash accounts (flatten nested structure)
        cash_accounts = self.financial_config.get('cash_accounts', {})
        cash_flat = self.flatten_nested_dict(cash_accounts, 'balance')
        for account, balance in sorted(cash_flat.items()):
            self.add_account_entry(self.cash_frame, self.cash_entries, account, balance)

        # Add new account button
        ttk.Button(
            self.cash_frame,
            text="+ Add Cash Account",
            command=lambda: self.add_new_account('cash')
        ).pack(pady=5)

        # Populate debt balances (flatten nested structure + credit cards)
        debt_balances = self.financial_config.get('debt_balances', {})
        debt_flat = self.flatten_nested_dict(debt_balances, 'balance')

        credit_cards = self.financial_config.get('credit_cards', {})
        credit_flat = self.flatten_nested_dict(credit_cards, 'balance')

        all_debt = {**debt_flat, **credit_flat}
        for account, balance in sorted(all_debt.items()):
            if balance > 0:  # Only show non-zero debts
                self.add_account_entry(self.debt_frame, self.debt_entries, account, balance)

        # Add new debt button
        ttk.Button(
            self.debt_frame,
            text="+ Add Debt Account",
            command=lambda: self.add_new_account('debt')
        ).pack(pady=5)

        # Populate recurring expenses (flatten nested structure)
        recurring = self.financial_config.get('recurring_expenses', {})
        recurring_flat = self.flatten_nested_dict(recurring, 'amount')
        for expense, amount in sorted(recurring_flat.items()):
            if amount > 0:  # Only show active expenses
                self.add_account_entry(self.recurring_frame, self.recurring_entries, expense, amount)

        # Add new recurring button
        ttk.Button(
            self.recurring_frame,
            text="+ Add Recurring Expense",
            command=lambda: self.add_new_account('recurring')
        ).pack(pady=5)

    def add_account_entry(self, parent, entry_dict, name, value):
        """Add an account entry field"""
        frame = ttk.Frame(parent)
        frame.pack(fill='x', pady=2)

        label = ttk.Label(frame, text=name, width=30)
        label.pack(side='left', padx=5)

        entry = ttk.Entry(frame, width=20)
        entry.insert(0, str(value))
        entry.pack(side='left', padx=5)

        ttk.Button(
            frame,
            text="🗑️",
            width=3,
            command=lambda: self.delete_account_entry(frame, entry_dict, name)
        ).pack(side='left')

        entry_dict[name] = entry

    def add_new_account(self, account_type):
        """Add a new account entry"""
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Add New {account_type.title()} Account")
        dialog.geometry("400x150")

        ttk.Label(dialog, text="Account Name:").pack(pady=10)
        name_entry = ttk.Entry(dialog, width=40)
        name_entry.pack(pady=5)

        ttk.Label(dialog, text="Balance/Amount:").pack(pady=10)
        value_entry = ttk.Entry(dialog, width=40)
        value_entry.pack(pady=5)

        def save_new_account():
            name = name_entry.get().strip()
            try:
                value = float(value_entry.get().strip())
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid number for balance/amount")
                return

            if not name:
                messagebox.showerror("Error", "Please enter an account name")
                return

            if account_type == 'cash':
                self.add_account_entry(self.cash_frame, self.cash_entries, name, value)
                self.financial_config.setdefault('cash_accounts', {})[name] = value
            elif account_type == 'debt':
                self.add_account_entry(self.debt_frame, self.debt_entries, name, value)
                self.financial_config.setdefault('debt_balances', {})[name] = value
            elif account_type == 'recurring':
                self.add_account_entry(self.recurring_frame, self.recurring_entries, name, value)
                self.financial_config.setdefault('recurring_expenses', {})[name] = value

            dialog.destroy()

        ttk.Button(dialog, text="Save", command=save_new_account, style='Success.TButton').pack(pady=10)

    def delete_account_entry(self, frame, entry_dict, name):
        """Delete an account entry"""
        if messagebox.askyesno("Confirm Delete", f"Delete {name}?"):
            frame.destroy()
            if name in entry_dict:
                del entry_dict[name]

    def update_dashboard_metrics(self):
        """Update dashboard metric cards"""
        try:
            snapshot = self.dashboard_data.get('current_snapshot', {})

            # Update cards
            self.update_card('cash', f"${snapshot.get('liquid_cash', 0):,.2f}")
            self.update_card('debt', f"${snapshot.get('total_debt', 0):,.2f}")
            self.update_card('networth', f"${snapshot.get('net_worth', 0):,.2f}")
            self.update_card('recurring', f"${snapshot.get('monthly_recurring', 0):,.2f}")

            # Calculate monthly income from config
            monthly_income = self.calculate_monthly_income()
            self.update_card('income', f"${monthly_income:,.2f}")

            # Emergency fund progress
            goals = self.dashboard_data.get('goals', [])
            emergency_goal = next((g for g in goals if 'Emergency' in g.get('name', '')), None)
            if emergency_goal:
                progress = emergency_goal.get('progress_percent', 0)
                self.update_card('emergency', f"{progress:.1f}%")

            # Update alerts
            self.update_alerts()

        except Exception as e:
            self.log_activity(f"Error updating dashboard: {str(e)}")

    def update_card(self, key, value):
        """Update a metric card value"""
        if key in self.dashboard_cards:
            self.dashboard_cards[key].value_label.configure(text=value)

    def update_alerts(self):
        """Update alerts display"""
        self.alerts_text.delete('1.0', tk.END)

        alerts = self.dashboard_data.get('alerts', [])
        if alerts:
            for alert in alerts:
                alert_type = alert.get('type', 'info')
                message = alert.get('message', '')
                self.alerts_text.insert(tk.END, f"[{alert_type.upper()}] {message}\n")
        else:
            self.alerts_text.insert(tk.END, "No alerts - everything looks good! ✓")

    def calculate_monthly_income(self):
        """Calculate total monthly income"""
        # This could be enhanced to read from config
        return 11877.0  # Placeholder - should read from config

    # ================== CSV Import ==================

    def browse_import_file(self):
        """Browse for CSV or Excel files to import (multiple selection supported)"""
        filenames = filedialog.askopenfilenames(
            title="Select CSV or Excel Files (multiple selection supported)",
            filetypes=[
                ("CSV and Excel files", "*.csv *.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        if filenames:
            self.import_files = list(filenames)
            # Show file count in entry
            if len(filenames) == 1:
                self.import_file_var.set(filenames[0])
            else:
                self.import_file_var.set(f"{len(filenames)} files selected")

            # Enable validate button
            self.import_btn.configure(state='normal')

    def clear_import_selection(self):
        """Clear file selection"""
        self.import_files = []
        self.import_file_var.set("")
        self.import_preview.delete('1.0', tk.END)
        self.import_preview.insert(tk.END, "No files selected. Click 'Browse...' to select files.\n")
        self.import_btn.configure(state='disabled')

    def read_file_data(self, filename):
        """Read data from CSV or Excel file"""
        file_path = Path(filename)
        file_ext = file_path.suffix.lower()

        if file_ext == '.csv':
            # Read CSV
            with open(filename, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames
                rows = [row for row in reader]
            return headers, rows

        elif file_ext in ['.xlsx', '.xls']:
            # Try to read Excel file
            try:
                import openpyxl
                from openpyxl import load_workbook

                wb = load_workbook(filename, read_only=True)
                ws = wb.active

                # Get headers from first row
                headers = [cell.value for cell in ws[1]]

                # Get data rows
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)

                wb.close()
                return headers, rows

            except ImportError:
                raise Exception("openpyxl library not installed. Install with: pip install openpyxl")
            except Exception as e:
                raise Exception(f"Error reading Excel file: {str(e)}")
        else:
            raise Exception(f"Unsupported file type: {file_ext}")

    def validate_files(self):
        """Validate all selected files"""
        if not self.import_files:
            messagebox.showwarning("No Files", "Please select at least one file first")
            return

        self.import_preview.delete('1.0', tk.END)
        self.import_preview.insert(tk.END, f"Validating {len(self.import_files)} file(s)...\n")
        self.import_preview.insert(tk.END, "="*60 + "\n\n")

        all_valid = True

        for file_idx, filename in enumerate(self.import_files, 1):
            file_path = Path(filename)
            self.import_preview.insert(tk.END, f"\n[{file_idx}/{len(self.import_files)}] {file_path.name}\n")
            self.import_preview.insert(tk.END, "-"*60 + "\n")

            try:
                headers, rows = self.read_file_data(filename)

                self.import_preview.insert(tk.END, f"✓ Type: {file_path.suffix.upper()[1:]}\n")
                self.import_preview.insert(tk.END, f"✓ Headers found: {', '.join([str(h) for h in headers if h])}\n")
                self.import_preview.insert(tk.END, f"✓ Rows: {len(rows)}\n")

                # Show first 2 rows as preview
                preview_count = min(2, len(rows))
                if preview_count > 0:
                    self.import_preview.insert(tk.END, f"\nPreview (first {preview_count} rows):\n")
                    for i, row in enumerate(rows[:preview_count], 1):
                        self.import_preview.insert(tk.END, f"\n  Row {i}:\n")
                        for key, value in list(row.items())[:5]:  # Show first 5 columns
                            if key:
                                self.import_preview.insert(tk.END, f"    {key}: {value}\n")
                        if len(row) > 5:
                            self.import_preview.insert(tk.END, f"    ... and {len(row)-5} more columns\n")

                # Validation checks
                required_fields = ['Date', 'Amount']
                missing = [f for f in required_fields if f not in headers]

                if missing:
                    self.import_preview.insert(tk.END, f"\n⚠️ Warning: Missing recommended fields: {', '.join(missing)}\n")
                    self.import_preview.insert(tk.END, "   File may still be importable but might need manual review.\n")
                else:
                    self.import_preview.insert(tk.END, f"\n✓ Validation passed!\n")

            except Exception as e:
                self.import_preview.insert(tk.END, f"\n❌ Error: {str(e)}\n")
                all_valid = False

        self.import_preview.insert(tk.END, "\n" + "="*60 + "\n")

        if all_valid:
            self.import_preview.insert(tk.END, f"\n✓ All {len(self.import_files)} file(s) validated successfully! Ready to import.\n")
            self.import_btn.configure(state='normal')
        else:
            self.import_preview.insert(tk.END, f"\n⚠️ Some files have validation issues. Review errors above.\n")
            self.import_btn.configure(state='disabled')

        self.import_preview.see(tk.END)

    def validate_csv(self):
        """Legacy single-file validation - redirects to new method"""
        self.validate_files()

    def convert_excel_to_csv(self, excel_file, csv_file):
        """Convert Excel file to CSV"""
        try:
            headers, rows = self.read_file_data(excel_file)

            with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                if headers and rows:
                    writer = csv.DictWriter(f, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(rows)

            return True
        except Exception as e:
            raise Exception(f"Excel to CSV conversion failed: {str(e)}")

    def import_csv(self):
        """Import and process all selected files (CSV and Excel)"""
        if not self.import_files:
            messagebox.showwarning("No Files", "Please select files to import first")
            return

        self.import_preview.insert(tk.END, "\n" + "="*60 + "\n")
        self.import_preview.insert(tk.END, f"Importing {len(self.import_files)} file(s)...\n")
        self.import_preview.insert(tk.END, "="*60 + "\n\n")

        imported_count = 0
        error_count = 0

        for file_idx, filename in enumerate(self.import_files, 1):
            file_path = Path(filename)
            self.import_preview.insert(tk.END, f"\n[{file_idx}/{len(self.import_files)}] {file_path.name}\n")
            self.import_preview.insert(tk.END, "-"*60 + "\n")

            try:
                dest_path = self.raw_path / "exports" / file_path.name
                dest_path.parent.mkdir(parents=True, exist_ok=True)

                # Handle Excel files - convert to CSV first
                if file_path.suffix.lower() in ['.xlsx', '.xls']:
                    self.import_preview.insert(tk.END, "Converting Excel to CSV...\n")
                    csv_dest_path = dest_path.with_suffix('.csv')

                    self.convert_excel_to_csv(filename, csv_dest_path)
                    self.import_preview.insert(tk.END, f"✓ Converted to: {csv_dest_path.name}\n")
                    dest_path = csv_dest_path

                else:
                    # CSV file - copy directly
                    source_path = Path(filename).resolve()
                    dest_path_resolved = dest_path.resolve()

                    if source_path != dest_path_resolved:
                        shutil.copy(filename, dest_path)
                        self.import_preview.insert(tk.END, f"✓ Copied to: {dest_path}\n")
                    else:
                        self.import_preview.insert(tk.END, f"✓ Already in correct location\n")

                # Run import script if it exists
                import_script = self.scripts_path / "import_rocket_money.py"
                if import_script.exists():
                    self.import_preview.insert(tk.END, "Running import script...\n")
                    result = subprocess.run(
                        [sys.executable, str(import_script)],
                        cwd=str(self.base_path),
                        capture_output=True,
                        text=True,
                        encoding='utf-8',
                        errors='replace',
                        timeout=60
                    )

                    if result.returncode == 0:
                        self.import_preview.insert(tk.END, "✓ Import successful!\n")
                        if result.stdout:
                            # Show first few lines of output
                            output_lines = result.stdout.split('\n')[:5]
                            for line in output_lines:
                                if line.strip():
                                    self.import_preview.insert(tk.END, f"  {line}\n")
                    else:
                        self.import_preview.insert(tk.END, f"⚠️ Import warnings:\n")
                        if result.stderr:
                            self.import_preview.insert(tk.END, f"  {result.stderr[:200]}\n")

                imported_count += 1
                self.log_activity(f"Imported: {file_path.name}")

            except Exception as e:
                error_count += 1
                self.import_preview.insert(tk.END, f"❌ Error: {str(e)}\n")
                self.log_activity(f"Import failed: {file_path.name} - {str(e)}")

        # Summary
        self.import_preview.insert(tk.END, "\n" + "="*60 + "\n")
        self.import_preview.insert(tk.END, f"\nImport Summary:\n")
        self.import_preview.insert(tk.END, f"  ✓ Successfully imported: {imported_count}\n")
        if error_count > 0:
            self.import_preview.insert(tk.END, f"  ❌ Failed: {error_count}\n")
        self.import_preview.insert(tk.END, "\n")
        self.import_preview.see(tk.END)

        if imported_count > 0:
            messagebox.showinfo("Import Complete",
                              f"Successfully imported {imported_count} file(s)!" +
                              (f"\n{error_count} file(s) failed." if error_count > 0 else ""))

            # Trigger rebuild if auto-rebuild is enabled
            if self.auto_rebuild_var.get():
                self.rebuild_dashboard()
        else:
            messagebox.showerror("Import Failed", "No files were imported successfully.")

    # ================== Save Updates ==================

    def save_all_updates(self):
        """Save all balance updates"""
        try:
            # Validate if enabled
            if self.validate_on_save_var.get():
                if not self.validate_changes():
                    return

            # Collect new values
            cash_updates = {}
            for name, entry in self.cash_entries.items():
                try:
                    value = float(entry.get())
                    cash_updates[name] = value
                except ValueError:
                    messagebox.showerror("Error", f"Invalid value for {name}")
                    return

            debt_updates = {}
            for name, entry in self.debt_entries.items():
                try:
                    value = float(entry.get())
                    debt_updates[name] = value
                except ValueError:
                    messagebox.showerror("Error", f"Invalid value for {name}")
                    return

            recurring_updates = {}
            for name, entry in self.recurring_entries.items():
                try:
                    value = float(entry.get())
                    recurring_updates[name] = value
                except ValueError:
                    messagebox.showerror("Error", f"Invalid value for {name}")
                    return

            # Update cash accounts in nested structure
            cash_accounts = self.financial_config.get('cash_accounts', {})
            for category, items in cash_accounts.items():
                if isinstance(items, dict) and 'balance' in items:
                    # Direct item
                    if category in cash_updates:
                        items['balance'] = cash_updates[category]
                else:
                    # Nested category
                    for item_key, item_val in items.items():
                        if isinstance(item_val, dict) and item_key in cash_updates:
                            item_val['balance'] = cash_updates[item_key]

            # Update debt balances in nested structure
            debt_balances = self.financial_config.get('debt_balances', {})
            for category, items in debt_balances.items():
                if isinstance(items, dict):
                    for item_key, item_val in items.items():
                        if isinstance(item_val, dict) and item_key in debt_updates:
                            item_val['balance'] = debt_updates[item_key]

            # Update credit cards
            credit_cards = self.financial_config.get('credit_cards', {})
            for card_key, card_val in credit_cards.items():
                if isinstance(card_val, dict) and card_key in debt_updates:
                    card_val['balance'] = debt_updates[card_key]

            # Update recurring expenses in nested structure
            recurring = self.financial_config.get('recurring_expenses', {})
            for category, items in recurring.items():
                if isinstance(items, dict):
                    for item_key, item_val in items.items():
                        if isinstance(item_val, dict) and item_key in recurring_updates:
                            item_val['amount'] = recurring_updates[item_key]

            # Update metadata
            self.financial_config['metadata']['last_updated'] = datetime.now().isoformat()

            # Save to file
            config_file = self.processed_path / "financial_config.json"
            config_file.parent.mkdir(parents=True, exist_ok=True)

            with open(config_file, 'w') as f:
                json.dump(self.financial_config, f, indent=2)

            self.log_activity("✓ All changes saved successfully")
            messagebox.showinfo("Success", "All changes saved successfully!")

            # Trigger rebuild if auto-rebuild is enabled
            if self.auto_rebuild_var.get():
                self.rebuild_dashboard()

        except Exception as e:
            messagebox.showerror("Error", f"Save failed: {str(e)}")
            self.log_activity(f"ERROR: Save failed - {str(e)}")

    def validate_changes(self):
        """Validate changes before saving"""
        if not self.alert_large_changes_var.get():
            return True

        warnings = []

        # Check cash accounts for large changes
        old_cash_flat = self.flatten_nested_dict(self.financial_config.get('cash_accounts', {}), 'balance')
        for name, entry in self.cash_entries.items():
            try:
                new_value = float(entry.get())
                old_value = old_cash_flat.get(name, 0)

                if old_value > 0:
                    change_pct = abs((new_value - old_value) / old_value) * 100
                    if change_pct > 20:
                        warnings.append(f"Cash account '{name}': {change_pct:.1f}% change (${old_value:.2f} → ${new_value:.2f})")
            except ValueError:
                pass

        # Check debt balances for large changes
        debt_flat = self.flatten_nested_dict(self.financial_config.get('debt_balances', {}), 'balance')
        credit_flat = self.flatten_nested_dict(self.financial_config.get('credit_cards', {}), 'balance')
        old_debt_flat = {**debt_flat, **credit_flat}

        for name, entry in self.debt_entries.items():
            try:
                new_value = float(entry.get())
                old_value = old_debt_flat.get(name, 0)

                if old_value > 0:
                    change_pct = abs((new_value - old_value) / old_value) * 100
                    if change_pct > 20:
                        warnings.append(f"Debt account '{name}': {change_pct:.1f}% change (${old_value:.2f} → ${new_value:.2f})")
            except ValueError:
                pass

        if warnings:
            warning_msg = "Large changes detected:\n\n" + "\n".join(warnings) + "\n\nContinue saving?"
            return messagebox.askyesno("Large Changes Detected", warning_msg)

        return True

    # ================== Dashboard Rebuild ==================

    def rebuild_dashboard(self):
        """Rebuild the financial dashboard"""
        self.set_status("Building...", "warning")
        self.log_activity("Starting dashboard rebuild...")

        def build_thread():
            try:
                build_script = self.base_path / "build_all.py"
                if not build_script.exists():
                    raise FileNotFoundError("build_all.py not found")

                result = subprocess.run(
                    [sys.executable, str(build_script)],
                    cwd=str(self.base_path),
                    capture_output=True,
                    text=True,
                    encoding='utf-8',
                    errors='replace',
                    timeout=300  # 5 minute timeout
                )

                if result.returncode == 0:
                    self.root.after(0, lambda: self.set_status("✓ Build Complete", "success"))
                    self.root.after(0, lambda: self.log_activity("✓ Dashboard rebuilt successfully"))
                    self.root.after(0, self.load_financial_data)
                else:
                    self.root.after(0, lambda: self.set_status("Build Failed", "danger"))
                    self.root.after(0, lambda: self.log_activity(f"ERROR: Build failed\n{result.stderr}"))

            except Exception as e:
                self.root.after(0, lambda: self.set_status("Build Error", "danger"))
                self.root.after(0, lambda: self.log_activity(f"ERROR: {str(e)}"))

        threading.Thread(target=build_thread, daemon=True).start()

    # ================== Snapshots ==================

    def save_snapshot(self):
        """Save monthly snapshot"""
        try:
            snapshot_script = self.scripts_path / "save_monthly_snapshot.py"
            if not snapshot_script.exists():
                messagebox.showerror("Error", "Snapshot script not found")
                return

            result = subprocess.run(
                [sys.executable, str(snapshot_script)],
                cwd=str(self.base_path),
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace'
            )

            if result.returncode == 0:
                self.log_activity("✓ Snapshot saved successfully")
                messagebox.showinfo("Success", "Monthly snapshot saved!")
                self.load_snapshots()
            else:
                messagebox.showerror("Error", f"Snapshot failed:\n{result.stderr}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save snapshot: {str(e)}")

    def load_snapshots(self):
        """Load available snapshots for comparison"""
        snapshots_path = self.archive_path / "snapshots"
        if not snapshots_path.exists():
            return

        snapshots = sorted([d.name for d in snapshots_path.iterdir() if d.is_dir()], reverse=True)

        if snapshots:
            self.compare_month1['values'] = snapshots
            self.compare_month2['values'] = snapshots

            if len(snapshots) >= 2:
                self.compare_month1.current(0)
                self.compare_month2.current(1)

    def run_comparison(self):
        """Run month-over-month comparison"""
        month1 = self.compare_month1.get()
        month2 = self.compare_month2.get()

        if not month1 or not month2:
            messagebox.showwarning("No Selection", "Please select two months to compare")
            return

        try:
            snapshot1_path = self.archive_path / "snapshots" / month1 / "dashboard_data.json"
            snapshot2_path = self.archive_path / "snapshots" / month2 / "dashboard_data.json"

            with open(snapshot1_path) as f:
                data1 = json.load(f)

            with open(snapshot2_path) as f:
                data2 = json.load(f)

            # Compare snapshots
            self.comparison_text.delete('1.0', tk.END)
            self.comparison_text.insert(tk.END, f"Comparison: {month1} vs {month2}\n")
            self.comparison_text.insert(tk.END, "=" * 60 + "\n\n")

            snap1 = data1.get('current_snapshot', {})
            snap2 = data2.get('current_snapshot', {})

            metrics = [
                ('Liquid Cash', 'liquid_cash'),
                ('Total Debt', 'total_debt'),
                ('Net Worth', 'net_worth'),
                ('Monthly Recurring', 'monthly_recurring')
            ]

            for label, key in metrics:
                val1 = snap1.get(key, 0)
                val2 = snap2.get(key, 0)
                diff = val1 - val2
                pct = (diff / val2 * 100) if val2 != 0 else 0

                arrow = "↑" if diff > 0 else "↓" if diff < 0 else "→"

                self.comparison_text.insert(tk.END, f"{label}:\n")
                self.comparison_text.insert(tk.END, f"  {month1}: ${val1:,.2f}\n")
                self.comparison_text.insert(tk.END, f"  {month2}: ${val2:,.2f}\n")
                self.comparison_text.insert(tk.END, f"  Change: {arrow} ${abs(diff):,.2f} ({pct:+.1f}%)\n\n")

        except Exception as e:
            messagebox.showerror("Error", f"Comparison failed: {str(e)}")

    # ================== Reports ==================

    def refresh_reports_list(self):
        """Refresh list of available reports"""
        reports_path = self.archive_path / "reports"
        if not reports_path.exists():
            return

        reports = sorted([f.name for f in reports_path.iterdir() if f.suffix == '.md'])
        self.report_selector['values'] = reports

        if reports:
            self.report_selector.current(0)
            self.load_selected_report()

    def load_selected_report(self, event=None):
        """Load selected report"""
        report_name = self.report_selector.get()
        if not report_name:
            return

        try:
            report_path = self.archive_path / "reports" / report_name
            with open(report_path, 'r') as f:
                content = f.read()

            self.report_text.delete('1.0', tk.END)
            self.report_text.insert(tk.END, content)

        except Exception as e:
            self.report_text.delete('1.0', tk.END)
            self.report_text.insert(tk.END, f"Error loading report: {str(e)}")

    # ================== Auto-Rebuild Watcher ==================

    def toggle_auto_rebuild(self):
        """Toggle auto-rebuild watcher"""
        if self.auto_rebuild_var.get():
            self.start_watcher()
        else:
            self.stop_watcher()

    def start_watcher(self):
        """Start file watcher for auto-rebuild"""
        if self.watcher_running:
            return

        self.watcher_running = True
        self.rebuild_status.configure(text="Status: Running", foreground='green')
        self.log_activity("▶️ Auto-rebuild watcher started")

        def watch_thread():
            last_modified = {}
            watch_path = self.processed_path

            while self.watcher_running:
                try:
                    for file in watch_path.glob("*.json"):
                        current_mtime = file.stat().st_mtime

                        if file.name not in last_modified:
                            last_modified[file.name] = current_mtime
                        elif current_mtime > last_modified[file.name]:
                            last_modified[file.name] = current_mtime
                            self.root.after(0, lambda f=file.name: self.log_activity(f"📝 Detected change: {f}"))
                            self.root.after(0, lambda: self.log_activity("🔄 Auto-rebuilding dashboard..."))
                            self.root.after(0, self.rebuild_dashboard)
                            time.sleep(5)  # Debounce

                    time.sleep(2)  # Check every 2 seconds

                except Exception as e:
                    self.root.after(0, lambda: self.log_activity(f"Watcher error: {str(e)}"))

        self.watcher_thread = threading.Thread(target=watch_thread, daemon=True)
        self.watcher_thread.start()

    def stop_watcher(self):
        """Stop file watcher"""
        if not self.watcher_running:
            return

        self.watcher_running = False
        self.rebuild_status.configure(text="Status: Stopped", foreground='red')
        self.log_activity("⏹️ Auto-rebuild watcher stopped")
        self.auto_rebuild_var.set(False)

    # ================== Validation ==================

    def run_validation(self):
        """Run data validation checks"""
        self.automation_log.insert(tk.END, "\n" + "="*60 + "\n")
        self.automation_log.insert(tk.END, "Running data validation...\n")
        self.automation_log.insert(tk.END, "="*60 + "\n\n")

        issues = []

        # Check for negative cash
        cash_flat = self.flatten_nested_dict(self.financial_config.get('cash_accounts', {}), 'balance')
        for account, balance in cash_flat.items():
            if balance < 0:
                issues.append(f"⚠️ Negative cash balance: {account} = ${balance:,.2f}")

        # Check for zero debts (might be paid off)
        debt_flat = self.flatten_nested_dict(self.financial_config.get('debt_balances', {}), 'balance')
        credit_flat = self.flatten_nested_dict(self.financial_config.get('credit_cards', {}), 'balance')
        all_debt = {**debt_flat, **credit_flat}

        for account, balance in all_debt.items():
            if balance == 0:
                issues.append(f"ℹ️ Zero debt balance (paid off?): {account}")

        # Check for very large recurring expenses
        recurring_flat = self.flatten_nested_dict(self.financial_config.get('recurring_expenses', {}), 'amount')
        for expense, amount in recurring_flat.items():
            if amount > 2000:
                issues.append(f"⚠️ Large recurring expense: {expense} = ${amount:,.2f}/month")

        # Check total cash vs total debt
        total_cash = sum(cash_flat.values())
        total_debt = sum(all_debt.values())

        if total_cash < total_debt * 0.01:  # Less than 1% of debt
            issues.append(f"⚠️ Very low cash reserves: ${total_cash:,.2f} vs ${total_debt:,.2f} debt")

        # Display results
        if issues:
            self.automation_log.insert(tk.END, f"Found {len(issues)} potential issues:\n\n")
            for issue in issues:
                self.automation_log.insert(tk.END, f"{issue}\n")
        else:
            self.automation_log.insert(tk.END, "✓ No issues found - data looks good!\n")

        self.automation_log.insert(tk.END, "\nValidation complete.\n")
        self.automation_log.see(tk.END)

    # ================== Utilities ==================

    def open_dashboard(self):
        """Open the dashboard HTML in browser"""
        dashboard_path = self.base_path / "financial_hub.html"
        if dashboard_path.exists():
            import webbrowser
            webbrowser.open(f"file://{dashboard_path.absolute()}")
        else:
            messagebox.showwarning("Not Found", "Dashboard HTML not found. Try rebuilding first.")

    def set_status(self, message, status_type="info"):
        """Update status indicator"""
        colors = {
            'success': self.colors['success'],
            'danger': self.colors['danger'],
            'warning': self.colors['warning'],
            'info': self.colors['info']
        }

        self.status_label.configure(
            text=f"● {message}",
            foreground=colors.get(status_type, self.colors['info'])
        )

    def log_activity(self, message):
        """Log activity to activity text widget"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.activity_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.activity_text.see(tk.END)

        # Also log to automation log if available
        try:
            self.automation_log.insert(tk.END, f"[{timestamp}] {message}\n")
            self.automation_log.see(tk.END)
        except:
            pass


def main():
    """Main entry point"""
    root = tk.Tk()
    app = FinancialManagerGUI(root)

    # Handle window close
    def on_closing():
        if app.watcher_running:
            app.stop_watcher()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    # Center window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')

    root.mainloop()


if __name__ == "__main__":
    main()
