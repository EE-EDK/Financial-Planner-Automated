#!/usr/bin/env python3
"""
Create a sample budget Excel file for testing
"""

from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment

# Create sample budget
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "November 2025"

# Headers
headers = [
    'Date', 'Description', 'Auto', 'Utilities', 'Groceries', 'Medical',
    'Entertainment', 'Savings', 'Deposit', 'Balance'
]

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Sample data
data = [
    ['11/23/2025', 'TSP & OPM deposits', 50.00, 348.00, 80.00, None, None, 200.00, 1708.00, 1708.00],
    ['12/1/2025', 'visa', -63.80, -298.50, -45.00, None, -15.99, -145.28, -552.58, 1155.42],
    ['12/5/2025', '', None, 120.00, 55.00, 45.00, None, None, None, 980.42],
    ['12/10/2025', 'salary', None, None, None, None, None, 500.00, 3500.00, 4480.42],
]

for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Totals row
totals_row = len(data) + 2
ws.cell(row=totals_row, column=1, value='Totals').font = Font(bold=True)

# Calculate totals for numeric columns
for col_idx in range(3, len(headers) + 1):
    total = sum(
        ws.cell(row=r, column=col_idx).value or 0
        for r in range(2, totals_row)
    )
    ws.cell(row=totals_row, column=col_idx, value=total).font = Font(bold=True)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save
output_dir = Path(__file__).parent.parent / "Archive" / "raw" / "exports"
output_dir.mkdir(parents=True, exist_ok=True)
output_file = output_dir / "Sample_Budget.xlsx"

wb.save(output_file)
print(f"✅ Created sample budget file: {output_file}")
print()
print("Run the import script:")
print(f"  python financial_docs/scripts/import_excel_budget.py {output_file}")
