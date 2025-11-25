#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Budget Transaction Importer
Imports transaction Excel exports with budget format and processes them for analysis

This script handles Excel files with the following format:
- Row 1: Column headers (Date, category columns, Balance)
- Data rows: Date in first column, amounts in category columns
- Last row: Totals row (skipped)

Example format:
| Date       | Auto  | Utilities | Groceries | ... | Balance |
|------------|-------|-----------|-----------|-----|---------|
| 11/23/2025 | 50.00 | 348.00    | 80.00     | ... | 1708.00 |
| 12/1/2025  | -63.80| -298.50   | -45.00    | ... | 1155.42 |

Usage:
    python scripts/import_excel_budget.py [path/to/budget.xlsx]

Output:
    - Updates Archive/raw/exports/AllTransactions.csv
    - Generates transaction summary
    - Shows category breakdown
"""

import sys
import io
import csv
import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from decimal import Decimal

# Fix Windows console encoding
if sys.platform == 'win32' and sys.stdout is not None:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

BASE_DIR = Path(__file__).parent.parent
ARCHIVE_DIR = BASE_DIR / "Archive"
PROCESSED_DIR = ARCHIVE_DIR / "processed"
EXPORTS_DIR = ARCHIVE_DIR / "raw" / "exports"
CONFIG_FILE = PROCESSED_DIR / "financial_config.json"

# Category mapping for standardization
CATEGORY_MAPPING = {
    'xtra car': 'Auto & Transport',
    'auto': 'Auto & Transport',
    'car ins.': 'Auto Insurance',
    'car ins': 'Auto Insurance',
    'cycle ins.': 'Insurance',
    'cycle ins': 'Insurance',
    'home insurance': 'Home Insurance',
    'property taxes': 'Property Tax',
    'utilities': 'Bills & Utilities',
    'medical': 'Healthcare & Medical',
    'x-mas': 'Gifts & Donations',
    'xmas': 'Gifts & Donations',
    'vacation': 'Travel',
    'b-day': 'Gifts & Donations',
    'bday': 'Gifts & Donations',
    'clothes': 'Shopping',
    'clothing': 'Shopping',
    'furniture': 'Home & Garden',
    'mower': 'Home & Garden',
    'interest': 'Interest Income',
    'extra rsp': 'Savings & Investments',
    'savings': 'Savings & Investments',
    'deposit': 'Income',
    'janet\'s': 'Personal',
    'janets': 'Personal',
}

# Columns to skip (not expense/income categories)
SKIP_COLUMNS = {'date', 'balance', 'description', 'discription', 'notes'}

def normalize_category(column_name):
    """Normalize category name for mapping"""
    normalized = column_name.lower().strip()
    return CATEGORY_MAPPING.get(normalized, column_name.title())

def parse_date(date_value):
    """Parse date from various formats"""
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')

    if not date_value:
        return None

    date_str = str(date_value).strip()
    if not date_str:
        return None

    # Try various date formats
    date_formats = [
        '%m/%d/%Y',
        '%Y-%m-%d',
        '%m/%d/%y',
        '%d/%m/%Y',
        '%Y/%m/%d',
        '%m-%d-%Y',
        '%d-%m-%Y',
    ]

    for fmt in date_formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue

    # If no format matches, return None
    print(f"⚠️  Could not parse date: {date_str}")
    return None

def parse_amount(value):
    """Parse amount from cell value"""
    if value is None or value == '':
        return None

    try:
        # Handle string values
        if isinstance(value, str):
            # Remove whitespace and common formatting
            value = value.strip().replace(',', '').replace('$', '')
            if not value or value == '-':
                return None

        amount = Decimal(str(value))
        # Return None for zero amounts
        if amount == 0:
            return None
        return amount
    except (ValueError, ArithmeticError):
        return None

def import_excel_budget(excel_file):
    """Import Excel budget file with category columns"""

    print(f"📥 Importing Excel budget from: {excel_file}")
    print()

    try:
        # Try using openpyxl first
        import openpyxl
        use_openpyxl = True
    except ImportError:
        # Fall back to pandas if available
        try:
            import pandas as pd
            use_openpyxl = False
        except ImportError:
            print("❌ Error: Need openpyxl or pandas to read Excel files")
            print("   Install with: pip install openpyxl")
            return []

    transactions = []
    categories = defaultdict(Decimal)
    monthly_totals = defaultdict(Decimal)

    if use_openpyxl:
        # Parse using openpyxl
        wb = openpyxl.load_workbook(excel_file, data_only=True)

        # Process each sheet
        for sheet_name in wb.sheetnames:
            print(f"📊 Processing sheet: {sheet_name}")
            ws = wb[sheet_name]

            # Get header row (first row)
            headers = []
            for cell in ws[1]:
                header = str(cell.value).strip() if cell.value else ''
                headers.append(header)

            if not headers or not headers[0]:
                print(f"⚠️  Skipping sheet {sheet_name}: No headers found")
                continue

            # Find date column
            date_col_idx = None
            for idx, header in enumerate(headers):
                if header.lower() in ['date', 'dt']:
                    date_col_idx = idx
                    break

            if date_col_idx is None:
                print(f"⚠️  Skipping sheet {sheet_name}: No date column found")
                continue

            # Process data rows (skip header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) <= date_col_idx:
                    continue

                # Get date
                date_value = row[date_col_idx]
                parsed_date = parse_date(date_value)

                if not parsed_date:
                    # Skip rows without valid dates
                    continue

                # Check if this is the totals row
                date_str = str(date_value).lower().strip()
                if 'total' in date_str:
                    continue

                # Get description if available
                description = ''
                if len(row) > date_col_idx + 1 and headers[date_col_idx + 1].lower() in ['description', 'discription', 'notes']:
                    description = str(row[date_col_idx + 1]) if row[date_col_idx + 1] else ''

                # Process each category column
                for col_idx, (header, value) in enumerate(zip(headers, row)):
                    if col_idx == date_col_idx:
                        continue

                    # Skip non-category columns
                    if header.lower() in SKIP_COLUMNS:
                        continue

                    if not header or not header.strip():
                        continue

                    # Parse amount
                    amount = parse_amount(value)
                    if amount is None:
                        continue

                    # Normalize category
                    category = normalize_category(header)

                    # Create transaction
                    txn = {
                        'date': parsed_date,
                        'description': description if description else f"{category} - {parsed_date}",
                        'amount': float(amount),
                        'category': category,
                        'account': 'Budget Import',
                        'notes': f"Imported from {excel_file.name}"
                    }

                    transactions.append(txn)

                    # Track categories (expenses are positive in output)
                    categories[category] += abs(amount)

                    # Monthly totals
                    month = parsed_date[:7]  # YYYY-MM
                    monthly_totals[month] += abs(amount)

    else:
        # Parse using pandas
        xls = pd.ExcelFile(excel_file)

        for sheet_name in xls.sheet_names:
            print(f"📊 Processing sheet: {sheet_name}")
            df = pd.read_excel(excel_file, sheet_name=sheet_name)

            # Find date column
            date_col = None
            for col in df.columns:
                if str(col).lower() in ['date', 'dt']:
                    date_col = col
                    break

            if date_col is None:
                print(f"⚠️  Skipping sheet {sheet_name}: No date column found")
                continue

            # Process each row
            for idx, row in df.iterrows():
                date_value = row[date_col]
                parsed_date = parse_date(date_value)

                if not parsed_date:
                    continue

                # Check if totals row
                date_str = str(date_value).lower().strip()
                if 'total' in date_str:
                    continue

                # Get description if available
                description = ''
                for col in df.columns:
                    if str(col).lower() in ['description', 'discription', 'notes']:
                        description = str(row[col]) if pd.notna(row[col]) else ''
                        break

                # Process each category column
                for col in df.columns:
                    if col == date_col or str(col).lower() in SKIP_COLUMNS:
                        continue

                    # Parse amount
                    amount = parse_amount(row[col])
                    if amount is None:
                        continue

                    # Normalize category
                    category = normalize_category(str(col))

                    # Create transaction
                    txn = {
                        'date': parsed_date,
                        'description': description if description else f"{category} - {parsed_date}",
                        'amount': float(amount),
                        'category': category,
                        'account': 'Budget Import',
                        'notes': f"Imported from {excel_file.name}"
                    }

                    transactions.append(txn)

                    # Track categories
                    categories[category] += abs(amount)

                    # Monthly totals
                    month = parsed_date[:7]
                    monthly_totals[month] += abs(amount)

    print(f"✅ Imported {len(transactions)} transactions")
    print()

    # Summary
    if transactions:
        print("=" * 70)
        print("📊 SPENDING SUMMARY")
        print("=" * 70)
        print()

        print("Categories:")
        sorted_cats = sorted(categories.items(), key=lambda x: x[1], reverse=True)
        for cat, amount in sorted_cats:
            print(f"  {cat:30s} ${amount:>10,.2f}")
        print()

        print("Monthly Totals:")
        sorted_months = sorted(monthly_totals.items())
        for month, amount in sorted_months:
            print(f"  {month}: ${amount:,.2f}")
        print()

    return transactions

def save_transactions(transactions, excel_file):
    """Save transactions to AllTransactions.csv"""

    if not transactions:
        print("⚠️  No transactions to save")
        return

    output_file = EXPORTS_DIR / "AllTransactions.csv"

    # Read existing transactions
    existing = []
    existing_keys = set()

    if output_file.exists():
        with open(output_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                existing.append(row)
                # Create key for deduplication
                key = (row['Date'], row['Description'], row['Amount'], row['Category'])
                existing_keys.add(key)

    # Add new transactions (avoid duplicates)
    new_count = 0
    for txn in transactions:
        key = (txn['date'], txn['description'], str(txn['amount']), txn['category'])
        if key not in existing_keys:
            existing.append({
                'Date': txn['date'],
                'Description': txn['description'],
                'Amount': txn['amount'],
                'Category': txn['category'],
                'Account': txn['account'],
                'Notes': txn.get('notes', '')
            })
            existing_keys.add(key)
            new_count += 1

    # Sort by date
    existing.sort(key=lambda x: x['Date'])

    # Save
    fieldnames = ['Date', 'Description', 'Amount', 'Category', 'Account', 'Notes']
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(existing)

    print(f"✅ Saved to: {output_file}")
    print(f"   Total transactions: {len(existing)}")
    print(f"   New transactions: {new_count}")
    print()

def main():
    """Main import function"""

    print()
    print("=" * 70)
    print("📊 EXCEL BUDGET TRANSACTION IMPORTER")
    print("=" * 70)
    print()

    # Get Excel file path
    if len(sys.argv) > 1:
        excel_file = Path(sys.argv[1])
    else:
        # Look for most recent Excel file in Archive/raw/exports
        excel_files = list(EXPORTS_DIR.glob("*.xlsx")) + list(EXPORTS_DIR.glob("*.xls"))
        # Exclude temporary files
        excel_files = [f for f in excel_files if not f.name.startswith('~$')]

        if excel_files:
            excel_file = max(excel_files, key=lambda x: x.stat().st_mtime)
            print(f"📁 Found recent Excel file: {excel_file.name}")
        else:
            print("❌ No Excel budget file found")
            print()
            print("Usage:")
            print("  python scripts/import_excel_budget.py path/to/budget.xlsx")
            print()
            print("Or place Excel file in Archive/raw/exports/ and run without arguments")
            return 1

    if not excel_file.exists():
        print(f"❌ File not found: {excel_file}")
        return 1

    # Import transactions
    transactions = import_excel_budget(excel_file)

    # Save to AllTransactions.csv
    if transactions:
        save_transactions(transactions, excel_file)

    print("=" * 70)
    print("✅ IMPORT COMPLETE")
    print("=" * 70)
    print()
    print("Next steps:")
    print("  1. Run: python build_all.py")
    print("  2. Open: financial_hub.html")
    print()

    return 0

if __name__ == '__main__':
    sys.exit(main())
