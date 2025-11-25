#!/usr/bin/env python3
"""
Create a realistic sample budget Excel file matching the user's actual format
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Budget 2025"

# Headers matching the user's actual format
headers = [
    'Date', 'Description', 'Xtra Car', 'Auto', 'X-mas', 'Vacation',
    'Car Ins.', 'Cycle Ins.', 'B-Day', 'Clothes', 'Home Insurance',
    'Property Taxes', 'Utilities', 'Medical', "Janet's", 'Interest',
    'Mower', 'Furniture', 'Extra RSP', 'Savings', 'Deposit', 'Balance'
]

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Sample data matching the screenshot format
data = [
    ['11/23/2025', 'discription if necessary', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    ['12/1/2025', 'TSP of 916 & OPM of 75', 50.00, 80.00, 100.00, 100.00, 140.00, 10.00, 70.00, 110.00, 100.00, 270.00, 348.00, 80.00, None, None, None, 50.00, 200.00, None, 1708.00, 1708.00],
    ['12/1/2025', 'visa', None, None, None, None, None, None, None, -63.80, None, None, -298.50, -45.00, None, None, None, None, -145.28, None, -552.58, 1155.42],
]

# Write data
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Totals row
totals_row = len(data) + 2
ws.cell(row=totals_row, column=1, value='Totals').font = Font(bold=True)

# Calculate totals for numeric columns (skip Date, Description, Balance)
for col_idx in range(3, len(headers)):  # Start from column 3 (Xtra Car)
    if headers[col_idx - 1] == 'Balance':  # Skip Balance column
        continue
    total = sum(
        ws.cell(row=r, column=col_idx).value or 0
        for r in range(2, totals_row)
    )
    if total != 0:
        ws.cell(row=totals_row, column=col_idx, value=total).font = Font(bold=True)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 20)
    ws.column_dimensions[column].width = adjusted_width

# Save
output_dir = Path(__file__).parent.parent / "Archive" / "raw" / "exports"
output_dir.mkdir(parents=True, exist_ok=True)
output_file = output_dir / "Realistic_Budget.xlsx"

wb.save(output_file)
print(f"✅ Created realistic budget file: {output_file}")
print()
print("Categories included:")
for i, header in enumerate(headers[2:], start=3):  # Skip Date and Description
    print(f"  {i-2}. {header}")
