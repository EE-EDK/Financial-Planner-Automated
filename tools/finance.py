#!/usr/bin/env python3
"""
Mom's Budget Manager
Simplified Excel and CSV budget tracking

Usage:
    python finance.py              # Run full analysis (auto-detects Excel/CSV in data/inputs/)

Configuration:
    - Edit tools/config.json to define your budget categories
    - Drop Excel (.xlsx, .xls) or CSV file in data/inputs/
    - Run the program to generate financial_report.html
"""

import json
import sys
import os
import csv
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("‚ùå Excel support not available. Install: pip install openpyxl")
    sys.exit(1)

# Directories
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent.parent

TOOLS_DIR = BASE_DIR / "tools"
DATA_DIR = BASE_DIR / "data"
INPUTS_DIR = DATA_DIR / "inputs"
CONFIG_FILE = TOOLS_DIR / "config.json"
REPORT_FILE = BASE_DIR / "financial_report.html"

# Create directories
DATA_DIR.mkdir(exist_ok=True)
INPUTS_DIR.mkdir(exist_ok=True)


class BudgetManager:
    def __init__(self):
        self.config = self.load_config()
        self.monthly_data = {}  # {month: {date: row_data}}

    def load_config(self):
        """Load configuration from config.json"""
        if not CONFIG_FILE.exists():
            print(f"‚ùå Config file not found: {CONFIG_FILE}")
            print("   Create config.json using budget_editor.html")
            sys.exit(1)

        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)

        # Validate config structure
        required_keys = ['expenses', 'savings_goals', 'accounts']
        for key in required_keys:
            if key not in config:
                print(f"‚ùå Config missing required key: {key}")
                sys.exit(1)

        return config

    def find_input_files(self):
        """Find Excel and CSV files in data/inputs/"""
        input_files = (
            list(INPUTS_DIR.glob('*.xlsx')) +
            list(INPUTS_DIR.glob('*.xls')) +
            list(INPUTS_DIR.glob('*.csv'))
        )
        input_files = [f for f in input_files if not f.name.startswith('~')]  # Ignore temp files
        return input_files

    def parse_date(self, date_value):
        """Parse date from Excel (handles MM/DD/YYYY and datetime objects)"""
        if isinstance(date_value, datetime):
            return date_value

        if isinstance(date_value, str):
            # Try MM/DD/YYYY format
            try:
                return datetime.strptime(date_value, '%m/%d/%Y')
            except:
                pass
            # Try other common formats
            for fmt in ['%Y-%m-%d', '%m-%d-%Y', '%d/%m/%Y']:
                try:
                    return datetime.strptime(date_value, fmt)
                except:
                    continue

        return None

    def process_excel_file(self, excel_file):
        """Process Excel file based on config.json structure"""
        print(f"\nüìä Processing: {excel_file.name}")

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # Read header row to map columns
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Find column indices by matching headers
        date_col = 0  # Column A (should be "Date")
        desc_col = 1  # Column B (should be description)

        # Build mappings by matching Excel headers to config categories
        expense_cols = {}
        savings_cols = {}
        account_cols = {}

        print(f"   Matching Excel headers to config categories...")

        for col_idx, header in enumerate(header_row):
            if not header or col_idx < 2:  # Skip Date and Description columns
                continue

            header_str = str(header).strip()

            # Try to match to expense categories
            if header_str in self.config['expenses']:
                expense_cols[header_str] = col_idx
                print(f"      Found expense: {header_str} in column {chr(65 + col_idx)}")

            # Try to match to savings goals
            elif header_str in self.config['savings_goals']:
                savings_cols[header_str] = col_idx
                print(f"      Found savings goal: {header_str} in column {chr(65 + col_idx)}")

            # Try to match to account tracking
            elif header_str in self.config['accounts']:
                account_cols[header_str] = col_idx
                print(f"      Found account: {header_str} in column {chr(65 + col_idx)}")

            # Warn about unmatched columns
            else:
                if header_str not in ['Deposit', 'Balance', '']:  # Ignore common extra columns
                    print(f"      ‚ö†Ô∏è  Column '{header_str}' in Excel not in config (will be ignored)")

        # Verify we found all configured categories
        missing_expenses = set(self.config['expenses'].keys()) - set(expense_cols.keys())
        missing_savings = set(self.config['savings_goals'].keys()) - set(savings_cols.keys())
        missing_accounts = set(self.config['accounts'].keys()) - set(account_cols.keys())

        if missing_expenses:
            print(f"      ‚ö†Ô∏è  Config expenses not in Excel: {', '.join(missing_expenses)}")
        if missing_savings:
            print(f"      ‚ö†Ô∏è  Config savings goals not in Excel: {', '.join(missing_savings)}")
        if missing_accounts:
            print(f"      ‚ö†Ô∏è  Config accounts not in Excel: {', '.join(missing_accounts)}")

        print(f"   ‚úÖ Mapped {len(expense_cols)} expenses, {len(savings_cols)} savings, {len(account_cols)} accounts")

        # Process data rows
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Check for Totals row
            if row[0] and str(row[0]).strip().lower() == 'totals':
                print(f"   Found Totals row, stopping")
                break

            # Skip empty rows
            if not row[0]:
                continue

            # Parse date
            trans_date = self.parse_date(row[date_col])
            if not trans_date:
                continue

            # Get description
            description = str(row[desc_col]) if len(row) > desc_col and row[desc_col] else ""

            # Extract all values
            row_data = {
                'date': trans_date,
                'description': description,
                'expenses': {},
                'savings_goals': {},
                'accounts': {}
            }

            # Extract expenses
            for category, col_idx in expense_cols.items():
                if col_idx < len(row) and row[col_idx]:
                    try:
                        value = float(row[col_idx])
                        row_data['expenses'][category] = value
                    except:
                        pass

            # Extract savings goals
            for category, col_idx in savings_cols.items():
                if col_idx < len(row) and row[col_idx]:
                    try:
                        value = float(row[col_idx])
                        row_data['savings_goals'][category] = value
                    except:
                        pass

            # Extract account tracking
            for category, col_idx in account_cols.items():
                if col_idx < len(row) and row[col_idx]:
                    try:
                        value = float(row[col_idx])
                        row_data['accounts'][category] = value
                    except:
                        pass

            # Group by month
            month_key = trans_date.strftime('%Y-%m')
            if month_key not in self.monthly_data:
                self.monthly_data[month_key] = []

            self.monthly_data[month_key].append(row_data)
            row_count += 1

        print(f"   ‚úÖ Processed {row_count} rows across {len(self.monthly_data)} months")
        return True

    def process_csv_file(self, csv_file):
        """Process CSV file based on config.json structure"""
        print(f"\nüìä Processing: {csv_file.name}")

        with open(csv_file, 'r', encoding='utf-8') as f:
            # Read header row
            csv_reader = csv.reader(f)
            header_row = next(csv_reader)

            # Find column indices by matching headers
            date_col = 0  # Column A (should be "Date")
            desc_col = 1  # Column B (should be description)

            # Build mappings by matching CSV headers to config categories
            expense_cols = {}
            savings_cols = {}
            account_cols = {}

            print(f"   Matching CSV headers to config categories...")

            for col_idx, header in enumerate(header_row):
                if not header or col_idx < 2:  # Skip Date and Description columns
                    continue

                header_str = str(header).strip()

                # Try to match to expense categories
                if header_str in self.config['expenses']:
                    expense_cols[header_str] = col_idx
                    print(f"      Found expense: {header_str} in column {col_idx}")

                # Try to match to savings goals
                elif header_str in self.config['savings_goals']:
                    savings_cols[header_str] = col_idx
                    print(f"      Found savings goal: {header_str} in column {col_idx}")

                # Try to match to account tracking
                elif header_str in self.config['accounts']:
                    account_cols[header_str] = col_idx
                    print(f"      Found account: {header_str} in column {col_idx}")

                # Warn about unmatched columns
                else:
                    if header_str not in ['Deposit', 'Balance', '']:
                        print(f"      ‚ö†Ô∏è  Column '{header_str}' in CSV not in config (will be ignored)")

            # Verify we found all configured categories
            missing_expenses = set(self.config['expenses'].keys()) - set(expense_cols.keys())
            missing_savings = set(self.config['savings_goals'].keys()) - set(savings_cols.keys())
            missing_accounts = set(self.config['accounts'].keys()) - set(account_cols.keys())

            if missing_expenses:
                print(f"      ‚ö†Ô∏è  Config expenses not in CSV: {', '.join(missing_expenses)}")
            if missing_savings:
                print(f"      ‚ö†Ô∏è  Config savings goals not in CSV: {', '.join(missing_savings)}")
            if missing_accounts:
                print(f"      ‚ö†Ô∏è  Config accounts not in CSV: {', '.join(missing_accounts)}")

            print(f"   ‚úÖ Mapped {len(expense_cols)} expenses, {len(savings_cols)} savings, {len(account_cols)} accounts")

            # Process data rows
            row_count = 0
            for row in csv_reader:
                # Check for Totals row
                if row[0] and str(row[0]).strip().lower() == 'totals':
                    print(f"   Found Totals row, stopping")
                    break

                # Skip empty rows
                if not row[0]:
                    continue

                # Parse date
                trans_date = self.parse_date(row[date_col])
                if not trans_date:
                    continue

                # Get description
                description = str(row[desc_col]) if len(row) > desc_col and row[desc_col] else ""

                # Extract all values
                row_data = {
                    'date': trans_date,
                    'description': description,
                    'expenses': {},
                    'savings_goals': {},
                    'accounts': {}
                }

                # Extract expenses
                for category, col_idx in expense_cols.items():
                    if col_idx < len(row) and row[col_idx]:
                        try:
                            value = float(row[col_idx])
                            row_data['expenses'][category] = value
                        except:
                            pass

                # Extract savings goals
                for category, col_idx in savings_cols.items():
                    if col_idx < len(row) and row[col_idx]:
                        try:
                            value = float(row[col_idx])
                            row_data['savings_goals'][category] = value
                        except:
                            pass

                # Extract account tracking
                for category, col_idx in account_cols.items():
                    if col_idx < len(row) and row[col_idx]:
                        try:
                            value = float(row[col_idx])
                            row_data['accounts'][category] = value
                        except:
                            pass

                # Group by month
                month_key = trans_date.strftime('%Y-%m')
                if month_key not in self.monthly_data:
                    self.monthly_data[month_key] = []

                self.monthly_data[month_key].append(row_data)
                row_count += 1

            print(f"   ‚úÖ Processed {row_count} rows across {len(self.monthly_data)} months")
            return True

    def calculate_monthly_totals(self):
        """Calculate totals for each month"""
        monthly_totals = {}

        for month, rows in self.monthly_data.items():
            totals = {
                'expenses': defaultdict(float),
                'savings_goals': defaultdict(float),
                'accounts': defaultdict(float),
                'total_expenses': 0,
                'total_savings': 0,
                'total_deposits': 0,
                'start_balance': 0,
                'end_balance': 0
            }

            # Sum up all categories
            for row in rows:
                # Expenses
                for cat, val in row['expenses'].items():
                    totals['expenses'][cat] += abs(val)
                    totals['total_expenses'] += abs(val)

                # Savings goals
                for cat, val in row['savings_goals'].items():
                    totals['savings_goals'][cat] += abs(val)
                    totals['total_savings'] += abs(val)

                # Track balance (last value in month)
                if 'Balance' in row['accounts']:
                    totals['end_balance'] = row['accounts']['Balance']

                # Track deposits
                if 'Deposits' in row['accounts']:
                    totals['total_deposits'] += abs(row['accounts']['Deposits'])

            # Set start balance (first row's balance or previous month's end)
            if rows and 'Balance' in rows[0]['accounts']:
                totals['start_balance'] = rows[0]['accounts']['Balance']

            monthly_totals[month] = totals

        return monthly_totals

    def generate_html_report(self):
        """Generate beautiful HTML report"""
        print("\nüìÑ Generating HTML report...")

        monthly_totals = self.calculate_monthly_totals()

        # Sort months
        sorted_months = sorted(self.monthly_data.keys(), reverse=True)

        # Generate HTML - continuing in next message due to length
        html_start = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Mom's Budget Report</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }

        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            overflow: hidden;
        }

        .header {
            background: linear-gradient(135deg, #0F1626 0%, #1a2332 100%);
            color: white;
            padding: 50px 40px;
            text-align: center;
        }

        .header h1 {
            font-size: 3em;
            margin-bottom: 10px;
            font-weight: 300;
            letter-spacing: 3px;
        }

        .header p {
            color: #AB987A;
            font-size: 1.1em;
            margin-top: 10px;
        }

        .content {
            padding: 40px;
        }

        .month-section {
            margin-bottom: 50px;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }

        .month-header {
            background: linear-gradient(135deg, #0F1626 0%, #1a2332 100%);
            color: white;
            padding: 20px 30px;
            font-size: 1.8em;
            font-weight: 300;
            letter-spacing: 2px;
        }

        .month-content {
            background: #f8f9fa;
            padding: 30px;
        }

        .category-section {
            margin-bottom: 30px;
        }

        .category-section h3 {
            color: #0F1626;
            margin-bottom: 15px;
            font-size: 1.4em;
            padding-bottom: 10px;
            border-bottom: 2px solid #AB987A;
        }

        .category-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 12px;
            margin-top: 15px;
        }

        .category-item {
            background: white;
            padding: 12px 18px;
            border-radius: 8px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-left: 3px solid #AB987A;
        }

        .category-item .name {
            font-weight: 500;
            color: #0F1626;
        }

        .category-item .value {
            font-weight: 600;
            color: #667eea;
        }

        .summary-cards {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 30px 0;
        }

        .summary-card {
            background: white;
            padding: 20px;
            border-radius: 12px;
            text-align: center;
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        }

        .summary-card .label {
            font-size: 0.9em;
            color: #666;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 10px;
        }

        .summary-card .value {
            font-size: 2em;
            font-weight: 600;
            color: #0F1626;
        }

        .summary-card.positive .value {
            color: #4caf50;
        }

        .summary-card.negative .value {
            color: #f44336;
        }

        .footer {
            text-align: center;
            padding: 30px;
            color: #666;
            background: #f8f9fa;
        }

        .no-data {
            text-align: center;
            padding: 60px 20px;
            color: #999;
        }

        .no-data h2 {
            font-size: 2em;
            margin-bottom: 20px;
        }

        .no-data p {
            font-size: 1.1em;
            line-height: 1.8;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>üí∞ BUDGET REPORT</h1>
            <p>Generated on """ + datetime.now().strftime('%B %d, %Y at %I:%M %p') + """</p>
        </div>

        <div class="content">
"""

        html_content = ""
        
        if not sorted_months:
            html_content = """
            <div class="no-data">
                <h2>No Data Found</h2>
                <p>Drop your Excel budget file in <code>data/inputs/</code> folder<br>
                and run the program again to generate your report.</p>
            </div>
"""
        else:
            for month in sorted_months:
                month_name = datetime.strptime(month, '%Y-%m').strftime('%B %Y')
                totals = monthly_totals[month]
                rows = self.monthly_data[month]

                month_html = f"""
            <div class="month-section">
                <div class="month-header">
                    üìÖ {month_name}
                </div>
                <div class="month-content">
                    <!-- Summary Cards -->
                    <div class="summary-cards">
                        <div class="summary-card">
                            <div class="label">Total Expenses</div>
                            <div class="value">${totals['total_expenses']:,.2f}</div>
                        </div>
                        <div class="summary-card">
                            <div class="label">Total Savings</div>
                            <div class="value">${totals['total_savings']:,.2f}</div>
                        </div>
                        <div class="summary-card">
                            <div class="label">Deposits</div>
                            <div class="value positive">${totals['total_deposits']:,.2f}</div>
                        </div>
                        <div class="summary-card">
                            <div class="label">End Balance</div>
                            <div class="value">${totals['end_balance']:,.2f}</div>
                        </div>
                    </div>

                    <!-- Expenses -->
                    <div class="category-section">
                        <h3>üí≥ Monthly Expenses</h3>
                        <div class="category-grid">
"""

                for category in self.config['expenses'].keys():
                    amount = totals['expenses'].get(category, 0)
                    if amount > 0:
                        month_html += f"""
                            <div class="category-item">
                                <span class="name">{category}</span>
                                <span class="value">${amount:,.2f}</span>
                            </div>
"""

                month_html += """
                        </div>
                    </div>

                    <!-- Savings Goals -->
                    <div class="category-section">
                        <h3>üéØ Savings Goals</h3>
                        <div class="category-grid">
"""

                for category in self.config['savings_goals'].keys():
                    amount = totals['savings_goals'].get(category, 0)
                    if amount > 0:
                        month_html += f"""
                            <div class="category-item">
                                <span class="name">{category}</span>
                                <span class="value">${amount:,.2f}</span>
                            </div>
"""

                month_html += """
                        </div>
                    </div>

                    <!-- Transaction Details -->
                    <div class="category-section">
                        <h3>üìù Transaction Log</h3>
                        <div style="background: white; padding: 15px; border-radius: 8px; max-height: 300px; overflow-y: auto;">
                            <table style="width: 100%; border-collapse: collapse;">
                                <thead>
                                    <tr style="background: #f8f9fa; text-align: left;">
                                        <th style="padding: 10px; border-bottom: 2px solid #e0e0e0;">Date</th>
                                        <th style="padding: 10px; border-bottom: 2px solid #e0e0e0;">Description</th>
                                        <th style="padding: 10px; border-bottom: 2px solid #e0e0e0; text-align: right;">Amount</th>
                                    </tr>
                                </thead>
                                <tbody>
"""

                for row in sorted(rows, key=lambda x: x['date']):
                    date_str = row['date'].strftime('%m/%d/%Y')
                    desc = row['description']

                    # Calculate total for this row
                    total_amount = (
                        sum(abs(v) for v in row['expenses'].values()) +
                        sum(abs(v) for v in row['savings_goals'].values())
                    )

                    if total_amount > 0:
                        month_html += f"""
                                    <tr>
                                        <td style="padding: 8px; border-bottom: 1px solid #f0f0f0;">{date_str}</td>
                                        <td style="padding: 8px; border-bottom: 1px solid #f0f0f0;">{desc}</td>
                                        <td style="padding: 8px; border-bottom: 1px solid #f0f0f0; text-align: right;">${total_amount:,.2f}</td>
                                    </tr>
"""

                month_html += """
                                </tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
"""
                html_content += month_html

        html_end = """
        </div>

        <div class="footer">
            Made with ‚ù§Ô∏è for Mom
        </div>
    </div>
</body>
</html>
"""

        # Write HTML file
        full_html = html_start + html_content + html_end
        with open(REPORT_FILE, 'w', encoding='utf-8') as f:
            f.write(full_html)

        print(f"‚úÖ Report saved: {REPORT_FILE}")
        return REPORT_FILE

    def run(self):
        """Main execution"""
        print("=" * 70)
        print("üöÄ MOM'S BUDGET MANAGER")
        print("=" * 70)

        # Find input files (Excel or CSV)
        input_files = self.find_input_files()

        if not input_files:
            print("\n‚ùå No input files found in data/inputs/")
            print("   Drop your budget Excel (.xlsx) or CSV file there and run again!")
            return False

        # Process each input file
        for input_file in input_files:
            try:
                # Detect file type and use appropriate processor
                if input_file.suffix.lower() == '.csv':
                    self.process_csv_file(input_file)
                else:  # .xlsx or .xls
                    self.process_excel_file(input_file)
            except Exception as e:
                print(f"‚ùå Error processing {input_file.name}: {e}")
                import traceback
                traceback.print_exc()
                continue

        # Generate report
        if self.monthly_data:
            self.generate_html_report()
            print("\n" + "=" * 70)
            print("‚úÖ COMPLETE!")
            print("=" * 70)
            print(f"\nüìä Open your report: {REPORT_FILE}")
            print(f"üí∞ Edit budget: budget_editor.html\n")
            return True
        else:
            print("\n‚ùå No data found in Excel files")
            return False


def main():
    """Main entry point"""
    manager = BudgetManager()
    success = manager.run()
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
